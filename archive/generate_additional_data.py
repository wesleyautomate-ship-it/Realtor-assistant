#!/usr/bin/env python3
"""
Additional Sample Data Generator for Dubai Real Estate RAG Application
Generates additional comprehensive datasets for testing at scale
"""

import pandas as pd
import numpy as np
import random
import json
import csv
from datetime import datetime, timedelta
import os
from faker import Faker
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

# Initialize Faker for realistic data
fake = Faker(['en_US', 'ar_SA'])

# Additional data constants
VENDOR_TYPES = ["Contractor", "Maintenance", "Cleaning", "Security", "Landscaping", "Plumbing", "Electrical", "HVAC", "Painting", "Carpentry"]
AGENCY_TYPES = ["Residential", "Commercial", "Luxury", "Investment", "Property Management", "Consulting"]
CLIENT_TYPES = ["Buyer", "Seller", "Investor", "Tenant", "Landlord", "Developer"]
PROPERTY_STATUSES = ["Available", "Under Contract", "Sold", "Rented", "Off Market", "Under Renovation", "Under Construction"]

def generate_vendors_data(num_vendors=150):
    """Generate vendor data"""
    vendors = []
    
    for i in range(num_vendors):
        vendor_type = random.choice(VENDOR_TYPES)
        
        vendor_data = {
            "id": i + 1,
            "vendor_name": fake.company(),
            "vendor_type": vendor_type,
            "contact_person": fake.name(),
            "email": fake.email(),
            "phone": fake.phone_number(),
            "address": fake.address(),
            "license_number": f"LIC-{random.randint(10000, 99999)}",
            "rating": round(random.uniform(1.0, 5.0), 1),
            "specializations": random.sample(["Residential", "Commercial", "Luxury", "Maintenance", "Renovation"], random.randint(1, 3)),
            "contract_start_date": fake.date_between(start_date='-2y', end_date='today').strftime("%Y-%m-%d"),
            "contract_end_date": fake.date_between(start_date='today', end_date='+3y').strftime("%Y-%m-%d"),
            "status": random.choice(["Active", "Inactive", "Suspended"]),
            "total_projects": random.randint(5, 200),
            "average_rating": round(random.uniform(3.0, 5.0), 1)
        }
        
        vendors.append(vendor_data)
    
    return vendors

def generate_agents_data(num_agents=100):
    """Generate agent data"""
    agents = []
    
    for i in range(num_agents):
        agency_type = random.choice(AGENCY_TYPES)
        
        agent_data = {
            "id": i + 1,
            "agent_id": f"AGT{str(i+1).zfill(4)}",
            "first_name": fake.first_name(),
            "last_name": fake.last_name(),
            "email": fake.email(),
            "phone": fake.phone_number(),
            "agency": fake.company(),
            "agency_type": agency_type,
            "license_number": f"RERA-{random.randint(100000, 999999)}",
            "specializations": random.sample(["Residential", "Commercial", "Luxury", "Off-plan", "Investment"], random.randint(1, 3)),
            "experience_years": random.randint(1, 20),
            "languages": random.sample(["English", "Arabic", "Hindi", "Urdu", "French", "German", "Chinese"], random.randint(1, 4)),
            "total_sales": random.randint(10, 500),
            "total_volume": random.randint(1000000, 50000000),
            "commission_rate": round(random.uniform(2.0, 5.0), 2),
            "performance_rating": round(random.uniform(1.0, 5.0), 1),
            "status": random.choice(["Active", "Inactive", "On Leave"]),
            "hire_date": fake.date_between(start_date='-5y', end_date='today').strftime("%Y-%m-%d")
        }
        
        agents.append(agent_data)
    
    return agents

def generate_clients_data(num_clients=300):
    """Generate client data"""
    clients = []
    
    for i in range(num_clients):
        client_type = random.choice(CLIENT_TYPES)
        
        client_data = {
            "id": i + 1,
            "client_id": f"CLT{str(i+1).zfill(4)}",
            "first_name": fake.first_name(),
            "last_name": fake.last_name(),
            "email": fake.email(),
            "phone": fake.phone_number(),
            "client_type": client_type,
            "nationality": random.choice(["UAE", "India", "Pakistan", "UK", "USA", "Canada", "Australia", "Germany", "France", "China"]),
            "preferred_areas": random.sample(["Downtown Dubai", "Dubai Marina", "Palm Jumeirah", "Business Bay", "Dubai Hills Estate"], random.randint(1, 3)),
            "budget_range": random.choice(["Under 1M", "1M-2M", "2M-5M", "5M-10M", "10M+"]),
            "property_preferences": random.sample(["Apartment", "Villa", "Townhouse", "Penthouse"], random.randint(1, 3)),
            "source": random.choice(["Website", "Referral", "Social Media", "Advertisement", "Direct Contact"]),
            "status": random.choice(["Active", "Inactive", "Converted", "Lost"]),
            "assigned_agent": random.randint(1, 100),
            "created_date": fake.date_between(start_date='-1y', end_date='today').strftime("%Y-%m-%d"),
            "last_contact": fake.date_between(start_date='-30d', end_date='today').strftime("%Y-%m-%d")
        }
        
        clients.append(client_data)
    
    return clients

def generate_listings_data(num_listings=400):
    """Generate listing data"""
    listings = []
    
    for i in range(num_listings):
        property_id = random.randint(1, 500)
        agent_id = random.randint(1, 100)
        
        listing_data = {
            "id": i + 1,
            "property_id": property_id,
            "agent_id": agent_id,
            "listing_title": f"Beautiful {random.choice(['Apartment', 'Villa', 'Townhouse'])} in {random.choice(['Downtown Dubai', 'Dubai Marina', 'Palm Jumeirah'])}",
            "listing_description": fake.text(max_nb_chars=500),
            "listing_price": random.randint(500000, 15000000),
            "listing_type": random.choice(["Sale", "Rent", "Lease"]),
            "status": random.choice(PROPERTY_STATUSES),
            "featured": random.choice([True, False]),
            "views_count": random.randint(0, 1000),
            "inquiries_count": random.randint(0, 50),
            "created_date": fake.date_between(start_date='-6m', end_date='today').strftime("%Y-%m-%d"),
            "last_updated": fake.date_between(start_date='-30d', end_date='today').strftime("%Y-%m-%d"),
            "expiry_date": fake.date_between(start_date='today', end_date='+6m').strftime("%Y-%m-%d"),
            "commission_rate": round(random.uniform(2.0, 5.0), 2)
        }
        
        listings.append(listing_data)
    
    return listings

def generate_property_amenities_data():
    """Generate property amenities mapping"""
    amenities_data = []
    
    for property_id in range(1, 501):  # For all 500 properties
        num_amenities = random.randint(3, 8)
        selected_amenities = random.sample([
            "Swimming Pool", "Gym", "Spa", "Tennis Court", "Basketball Court",
            "Children's Playground", "BBQ Area", "Garden", "Balcony", "Terrace",
            "Parking", "Security", "Concierge", "24/7 Maintenance", "Pet Friendly",
            "Central AC", "Furnished", "Built-in Wardrobes", "Modern Kitchen",
            "Walk-in Closet", "Study Room", "Maid's Room", "Driver's Room"
        ], num_amenities)
        
        for amenity in selected_amenities:
            amenity_data = {
                "property_id": property_id,
                "amenity_name": amenity,
                "amenity_type": random.choice(["Basic", "Premium", "Luxury"]),
                "included_in_rent": random.choice([True, False]),
                "additional_cost": random.randint(0, 5000) if random.choice([True, False]) else 0
            }
            amenities_data.append(amenity_data)
    
    return amenities_data

def generate_property_images_data():
    """Generate property images data"""
    images_data = []
    
    for property_id in range(1, 501):  # For all 500 properties
        num_images = random.randint(5, 15)
        
        for i in range(num_images):
            image_data = {
                "id": len(images_data) + 1,
                "property_id": property_id,
                "image_url": f"https://example.com/properties/{property_id}/image_{i+1}.jpg",
                "image_type": random.choice(["Exterior", "Interior", "Kitchen", "Bathroom", "Bedroom", "Living Room", "Balcony", "Amenities"]),
                "is_primary": i == 0,  # First image is primary
                "upload_date": fake.date_between(start_date='-6m', end_date='today').strftime("%Y-%m-%d"),
                "file_size": random.randint(500000, 5000000),  # 500KB to 5MB
                "resolution": random.choice(["1920x1080", "2560x1440", "3840x2160"])
            }
            images_data.append(image_data)
    
    return images_data

def generate_market_reports_data(num_reports=50):
    """Generate market reports data"""
    reports = []
    
    for i in range(num_reports):
        report_data = {
            "id": i + 1,
            "report_title": f"Market Report - {random.choice(['Q1', 'Q2', 'Q3', 'Q4'])} {random.randint(2022, 2024)}",
            "report_type": random.choice(["Quarterly", "Monthly", "Annual", "Special"]),
            "neighborhood": random.choice(["Downtown Dubai", "Dubai Marina", "Palm Jumeirah", "Business Bay", "Dubai Hills Estate"]),
            "property_type": random.choice(["Residential", "Commercial", "Luxury", "Investment"]),
            "report_date": fake.date_between(start_date='-2y', end_date='today').strftime("%Y-%m-%d"),
            "author": fake.name(),
            "summary": fake.text(max_nb_chars=300),
            "key_findings": random.randint(3, 8),
            "price_trend": random.choice(["Increasing", "Decreasing", "Stable", "Volatile"]),
            "market_sentiment": random.choice(["Positive", "Negative", "Neutral", "Optimistic", "Cautious"]),
            "recommendations": random.randint(2, 6),
            "file_path": f"reports/market_report_{i+1}.pdf"
        }
        
        reports.append(report_data)
    
    return reports

def create_additional_csv_files():
    """Create additional CSV files"""
    print("Generating additional CSV files...")
    
    # Vendors CSV
    vendors = generate_vendors_data(150)
    df_vendors = pd.DataFrame(vendors)
    df_vendors.to_csv('data/vendors.csv', index=False)
    print(f"Created vendors.csv with {len(vendors)} records")
    
    # Agents CSV
    agents = generate_agents_data(100)
    df_agents = pd.DataFrame(agents)
    df_agents.to_csv('data/agents.csv', index=False)
    print(f"Created agents.csv with {len(agents)} records")
    
    # Clients CSV
    clients = generate_clients_data(300)
    df_clients = pd.DataFrame(clients)
    df_clients.to_csv('data/clients.csv', index=False)
    print(f"Created clients.csv with {len(clients)} records")
    
    # Listings CSV
    listings = generate_listings_data(400)
    df_listings = pd.DataFrame(listings)
    df_listings.to_csv('data/listings.csv', index=False)
    print(f"Created listings.csv with {len(listings)} records")
    
    # Property Amenities CSV
    amenities = generate_property_amenities_data()
    df_amenities = pd.DataFrame(amenities)
    df_amenities.to_csv('data/property_amenities.csv', index=False)
    print(f"Created property_amenities.csv with {len(amenities)} records")
    
    # Property Images CSV
    images = generate_property_images_data()
    df_images = pd.DataFrame(images)
    df_images.to_csv('data/property_images.csv', index=False)
    print(f"Created property_images.csv with {len(images)} records")
    
    # Market Reports CSV
    reports = generate_market_reports_data(50)
    df_reports = pd.DataFrame(reports)
    df_reports.to_csv('data/market_reports.csv', index=False)
    print(f"Created market_reports.csv with {len(reports)} records")

def create_comprehensive_excel_file():
    """Create a comprehensive Excel file with multiple sheets"""
    print("Generating comprehensive Excel file...")
    
    wb = Workbook()
    
    # Remove default sheet
    wb.remove(wb.active)
    
    # Create sheets for different data types
    sheets_data = {
        "Agents": generate_agents_data(50),
        "Clients": generate_clients_data(100),
        "Vendors": generate_vendors_data(75),
        "Listings": generate_listings_data(100)
    }
    
    for sheet_name, data in sheets_data.items():
        ws = wb.create_sheet(sheet_name)
        
        if data:
            # Add headers
            headers = list(data[0].keys())
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
            
            # Add data
            for row, record in enumerate(data, 2):
                for col, header in enumerate(headers, 1):
                    value = record[header]
                    # Convert lists to strings for Excel compatibility
                    if isinstance(value, list):
                        value = ", ".join(str(item) for item in value)
                    ws.cell(row=row, column=col, value=value)
    
    wb.save('data/comprehensive_real_estate_data.xlsx')
    print("Created comprehensive_real_estate_data.xlsx with multiple sheets")

def create_json_datasets():
    """Create additional JSON datasets"""
    print("Generating additional JSON files...")
    
    # Market trends data
    market_trends = []
    for month in range(1, 13):
        for neighborhood in ["Downtown Dubai", "Dubai Marina", "Palm Jumeirah", "Business Bay"]:
            trend_data = {
                "month": month,
                "year": 2024,
                "neighborhood": neighborhood,
                "avg_price_per_sqft": random.randint(800, 3000),
                "price_change": round(random.uniform(-10.0, 15.0), 2),
                "transaction_volume": random.randint(10, 200),
                "days_on_market": random.randint(15, 120),
                "supply_level": random.choice(["Low", "Medium", "High"]),
                "demand_level": random.choice(["Low", "Medium", "High"])
            }
            market_trends.append(trend_data)
    
    with open('data/market_trends.json', 'w') as f:
        json.dump(market_trends, f, indent=2)
    print("Created market_trends.json")
    
    # Property analytics data
    property_analytics = []
    for property_id in range(1, 101):  # First 100 properties
        analytics_data = {
            "property_id": property_id,
            "views_last_30_days": random.randint(0, 500),
            "inquiries_last_30_days": random.randint(0, 20),
            "favorites_count": random.randint(0, 50),
            "shares_count": random.randint(0, 30),
            "avg_time_on_page": random.randint(30, 300),
            "bounce_rate": round(random.uniform(0.1, 0.8), 2),
            "conversion_rate": round(random.uniform(0.01, 0.1), 3),
            "price_history": [
                {"date": "2024-01-01", "price": random.randint(500000, 5000000)},
                {"date": "2024-02-01", "price": random.randint(500000, 5000000)},
                {"date": "2024-03-01", "price": random.randint(500000, 5000000)}
            ]
        }
        property_analytics.append(analytics_data)
    
    with open('data/property_analytics.json', 'w') as f:
        json.dump(property_analytics, f, indent=2)
    print("Created property_analytics.json")

def main():
    """Main function to generate additional sample data"""
    print("Starting additional sample data generation...")
    
    # Create data directory if it doesn't exist
    os.makedirs('data', exist_ok=True)
    
    # Generate additional data files
    create_additional_csv_files()
    create_comprehensive_excel_file()
    create_json_datasets()
    
    print("\nAdditional sample data generation completed!")
    print("Generated additional files:")
    print("- CSV files: vendors.csv, agents.csv, clients.csv, listings.csv, property_amenities.csv, property_images.csv, market_reports.csv")
    print("- Excel file: comprehensive_real_estate_data.xlsx")
    print("- JSON files: market_trends.json, property_analytics.json")

if __name__ == "__main__":
    main()
