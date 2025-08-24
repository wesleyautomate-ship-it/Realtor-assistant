"""
Data Ingestion Layer

Handles data ingestion from multiple sources including PDF, CSV, Excel, and web data.
"""

import pandas as pd
import PyPDF2
import openpyxl
import requests
from pathlib import Path
from typing import Dict, List, Any, Union, Optional
import logging
import json
from datetime import datetime

class DataIngestion:
    """Handles data ingestion from multiple sources"""
    
    def __init__(self, config: Dict[str, Any]):
        self.config = config
        self.logger = logging.getLogger(__name__)
        
    def ingest_pdf(self, file_path: str) -> Optional[Dict[str, Any]]:
        """Extract data from PDF files"""
        try:
            with open(file_path, 'rb') as file:
                pdf_reader = PyPDF2.PdfReader(file)
                text_content = ""
                for page in pdf_reader.pages:
                    text_content += page.extract_text()
                
                return {
                    'source': 'pdf',
                    'file_path': file_path,
                    'content': text_content,
                    'metadata': {
                        'pages': len(pdf_reader.pages),
                        'file_size': Path(file_path).stat().st_size,
                        'extracted_at': datetime.now().isoformat()
                    }
                }
        except Exception as e:
            self.logger.error(f"Error processing PDF {file_path}: {e}")
            return None
    
    def ingest_csv(self, file_path: str) -> Optional[Dict[str, Any]]:
        """Extract data from CSV files"""
        try:
            df = pd.read_csv(file_path)
            return {
                'source': 'csv',
                'file_path': file_path,
                'data': df.to_dict('records'),
                'metadata': {
                    'rows': len(df),
                    'columns': list(df.columns),
                    'data_types': df.dtypes.to_dict(),
                    'extracted_at': datetime.now().isoformat()
                }
            }
        except Exception as e:
            self.logger.error(f"Error processing CSV {file_path}: {e}")
            return None
    
    def ingest_excel(self, file_path: str) -> Optional[Dict[str, Any]]:
        """Extract data from Excel files"""
        try:
            workbook = openpyxl.load_workbook(file_path, data_only=True)
            sheets_data = {}
            
            for sheet_name in workbook.sheetnames:
                sheet = workbook[sheet_name]
                data = []
                headers = []
                
                for row in sheet.iter_rows(values_only=True):
                    if not headers:
                        headers = row
                    else:
                        data.append(dict(zip(headers, row)))
                
                sheets_data[sheet_name] = {
                    'headers': headers,
                    'data': data
                }
            
            return {
                'source': 'excel',
                'file_path': file_path,
                'sheets': sheets_data,
                'metadata': {
                    'sheets': list(workbook.sheetnames),
                    'total_rows': sum(len(sheet['data']) for sheet in sheets_data.values()),
                    'extracted_at': datetime.now().isoformat()
                }
            }
        except Exception as e:
            self.logger.error(f"Error processing Excel {file_path}: {e}")
            return None
    
    def ingest_web_data(self, url: str, scraper_type: str) -> Optional[Dict[str, Any]]:
        """Extract data from web sources"""
        try:
            headers = self.config.get('web_headers', {
                'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36'
            })
            
            response = requests.get(url, headers=headers, timeout=30)
            response.raise_for_status()
            
            return {
                'source': 'web',
                'url': url,
                'scraper_type': scraper_type,
                'content': response.text,
                'metadata': {
                    'status_code': response.status_code,
                    'content_length': len(response.text),
                    'extracted_at': datetime.now().isoformat()
                }
            }
        except Exception as e:
            self.logger.error(f"Error scraping {url}: {e}")
            return None
    
    def ingest_json(self, file_path: str) -> Optional[Dict[str, Any]]:
        """Extract data from JSON files"""
        try:
            with open(file_path, 'r', encoding='utf-8') as file:
                data = json.load(file)
                
            return {
                'source': 'json',
                'file_path': file_path,
                'data': data,
                'metadata': {
                    'file_size': Path(file_path).stat().st_size,
                    'extracted_at': datetime.now().isoformat()
                }
            }
        except Exception as e:
            self.logger.error(f"Error processing JSON {file_path}: {e}")
            return None
    
    def ingest_directory(self, directory_path: str, file_types: List[str] = None) -> List[Dict[str, Any]]:
        """Ingest all files of specified types from a directory"""
        if file_types is None:
            file_types = ['csv', 'xlsx', 'json', 'pdf']
        
        directory = Path(directory_path)
        ingested_files = []
        
        for file_path in directory.rglob('*'):
            if file_path.is_file() and file_path.suffix.lower()[1:] in file_types:
                try:
                    if file_path.suffix.lower() == '.csv':
                        result = self.ingest_csv(str(file_path))
                    elif file_path.suffix.lower() in ['.xlsx', '.xls']:
                        result = self.ingest_excel(str(file_path))
                    elif file_path.suffix.lower() == '.json':
                        result = self.ingest_json(str(file_path))
                    elif file_path.suffix.lower() == '.pdf':
                        result = self.ingest_pdf(str(file_path))
                    else:
                        continue
                    
                    if result:
                        ingested_files.append(result)
                        
                except Exception as e:
                    self.logger.error(f"Error processing {file_path}: {e}")
                    continue
        
        return ingested_files
    
    def validate_ingestion(self, ingested_data: Dict[str, Any]) -> Dict[str, Any]:
        """Validate ingested data quality"""
        validation_result = {
            'is_valid': True,
            'errors': [],
            'warnings': [],
            'data_quality_score': 0.0
        }
        
        if not ingested_data:
            validation_result['is_valid'] = False
            validation_result['errors'].append("No data ingested")
            return validation_result
        
        # Check for required fields based on source type
        source = ingested_data.get('source')
        
        if source == 'csv':
            if 'data' not in ingested_data or not ingested_data['data']:
                validation_result['is_valid'] = False
                validation_result['errors'].append("No data found in CSV")
            
            if 'metadata' in ingested_data:
                rows = ingested_data['metadata'].get('rows', 0)
                if rows == 0:
                    validation_result['warnings'].append("CSV file is empty")
                elif rows < 10:
                    validation_result['warnings'].append("CSV file has very few rows")
        
        elif source == 'excel':
            if 'sheets' not in ingested_data or not ingested_data['sheets']:
                validation_result['is_valid'] = False
                validation_result['errors'].append("No sheets found in Excel file")
        
        elif source == 'web':
            if 'content' not in ingested_data or not ingested_data['content']:
                validation_result['is_valid'] = False
                validation_result['errors'].append("No content extracted from web page")
        
        # Calculate data quality score
        if validation_result['is_valid']:
            validation_result['data_quality_score'] = 1.0 - (len(validation_result['warnings']) * 0.1)
        
        return validation_result
